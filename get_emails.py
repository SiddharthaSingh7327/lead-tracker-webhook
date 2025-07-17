import os
import msal
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta, timezone
import google.generativeai as genai
import json
import re
import pickle
import hashlib
from urllib.parse import quote
import logging
from typing import Dict, List, Optional, Set, Tuple
from dataclasses import dataclass
from pathlib import Path
import time
from functools import wraps
import msal
import json


# ---------------- Configuration ----------------
@dataclass
class Config:
    CLIENT_ID: str = "91d3f9fe-f30d-4409-85fa-fa4a7c24c047"
    TENANT_ID: str = "64a9da10-e764-406f-a749-552dade47aa9"
    GEMINI_API_KEY: str = "AIzaSyCtecm-I_JzMVNtQHsAfzRykn1XbKwuPXU"
    AUTHORITY: str = f"https://login.microsoftonline.com/64a9da10-e764-406f-a749-552dade47aa9"
    SCOPE: List[str] = None
    
    def __post_init__(self):
        if self.SCOPE is None:
            self.SCOPE = [
                "https://graph.microsoft.com/Mail.Read",
                "https://graph.microsoft.com/Calendars.ReadWrite",
                "https://graph.microsoft.com/Files.ReadWrite",
                "https://graph.microsoft.com/User.Read"
            ]


def get_token_with_cache(client_id, tenant_id, scopes):
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    app = msal.PublicClientApplication(client_id, authority=authority)

    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(scopes, account=accounts[0])
        if result:
            return result["access_token"]

    # Fallback to device flow
    flow = app.initiate_device_flow(scopes=scopes)
    print(flow["message"])  # Show login instructions
    result = app.acquire_token_by_device_flow(flow)

    if "access_token" in result:
        return result["access_token"]
    else:
        raise Exception(f"Auth failed: {result.get('error_description')}")
    

# ---------------- Logging Setup ----------------


def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('lead_tracker.log'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ---------------- Error Handling Decorator ----------------
def handle_exceptions(func):
    """Decorator to handle exceptions gracefully"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logger.error(f"Error in {func.__name__}: {str(e)}")
            return None
    return wrapper

# ---------------- Data Models ----------------
@dataclass
class EmailData:
    id: str
    sender_name: str
    sender_email: str
    subject: str
    received: str
    body_preview: str
    parsed_meeting: Optional[Dict] = None
    calendar_event_id: Optional[str] = None

@dataclass
class EventData:
    id: str
    subject: str
    start: str
    end: str
    location: str
    attendees: str
    body: str

@dataclass
class OpportunityData:
    id: str
    contact_name: str
    company: str
    email: str
    phone: str
    opportunity_title: str
    lead_status: str
    notes: str
    last_updated: str
    priority: str = "Medium"
    source: str = "Email"

# ---------------- File Management ----------------------------- #
class FileManager:
    """Handles file operations with better error handling"""
    
    def __init__(self, base_path: str = "."):
        self.base_path = Path(base_path)
        self.base_path.mkdir(exist_ok=True)
        
    def load_pickle(self, filename: str) -> Set:
        """Load pickle file with error handling"""
        filepath = self.base_path / filename
        if filepath.exists():
            try:
                with open(filepath, 'rb') as f:
                    return pickle.load(f)
            except Exception as e:
                logger.error(f"Error loading {filename}: {e}")
        return set()
    
    def save_pickle(self, data: Set, filename: str) -> bool:
        """Save pickle file with error handling"""
        filepath = self.base_path / filename
        try:
            with open(filepath, 'wb') as f:
                pickle.dump(data, f)
            return True
        except Exception as e:
            logger.error(f"Error saving {filename}: {e}")
            return False
    
    def load_json(self, filename: str) -> Dict:
        """Load JSON file with error handling"""
        filepath = self.base_path / filename
        if filepath.exists():
            try:
                with open(filepath, 'r') as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Error loading {filename}: {e}")
        return {}
    
    def save_json(self, data: Dict, filename: str) -> bool:
        """Save JSON file with error handling"""
        filepath = self.base_path / filename
        try:
            with open(filepath, 'w') as f:
                json.dump(data, f, indent=2)
            return True
        except Exception as e:
            logger.error(f"Error saving {filename}: {e}")
            return False

# ---------------- Enhanced Gemini Integration ----------------
class GeminiParser:
    """Enhanced Gemini integration with better error handling and retry logic"""
    
    def __init__(self, api_key: str):
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel("models/gemini-1.5-flash")
        
    def parse_email_with_retry(self, email_body: str, email_subject: str = "", max_retries: int = 3) -> Optional[Dict]:
        """Parse email with retry logic"""
        for attempt in range(max_retries):
            try:
                result = self._parse_email(email_body, email_subject)
                if result:
                    return result
            except Exception as e:
                logger.warning(f"Attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
        
        logger.error(f"Failed to parse email after {max_retries} attempts")
        return None
    
    def _parse_email(self, email_body: str, email_subject: str = "") -> Optional[Dict]:
        """Internal email parsing method"""
        prompt = f"""
        Extract structured meeting details from this email. Respond in **only valid JSON** format.

        Email Subject: {email_subject}
        Email Body: {email_body[:1000]}  # Truncate for token limits

        Extract meeting information if present. Convert relative dates to actual dates based on today: {datetime.now().strftime('%Y-%m-%d')}.

        Required JSON format:
        {{
          "has_meeting": true/false,
          "subject": "string or null",
          "date": "YYYY-MM-DD or null",
          "start_time": "HH:MM or null",
          "end_time": "HH:MM or null",
          "participants": ["email1@example.com"] or [],
          "location": "string or null",
          "action_items": "string or null",
          "deadline": "YYYY-MM-DD or null",
          "meeting_type": "Teams/Zoom/In-person/Call or null",
          "priority": "High/Medium/Low or null"
        }}
        """

        response = self.model.generate_content(prompt)
        raw_output = response.text.strip()
        
        # Clean up response
        raw_output = self._clean_json_response(raw_output)
        
        # Parse JSON
        parsed_data = json.loads(raw_output)
        
        # Validate required fields
        if not isinstance(parsed_data, dict):
            raise ValueError("Response is not a valid JSON object")
            
        return parsed_data
    
    def _clean_json_response(self, raw_output: str) -> str:
        """Clean up Gemini response to ensure valid JSON"""
        # Remove markdown formatting
        if raw_output.startswith("```json"):
            raw_output = raw_output[7:]
        if raw_output.endswith("```"):
            raw_output = raw_output[:-3]
        
        # Remove any trailing text after JSON
        lines = raw_output.split('\n')
        json_lines = []
        bracket_count = 0
        
        for line in lines:
            json_lines.append(line)
            bracket_count += line.count('{') - line.count('}')
            if bracket_count == 0 and line.strip().endswith('}'):
                break
        
        return '\n'.join(json_lines).strip()

# ---------------- Calendar Management ----------------
class CalendarManager:
    """Enhanced calendar management with better duplicate detection"""
    
    def __init__(self, headers: Dict, file_manager: FileManager):
        self.headers = headers
        self.file_manager = file_manager
        self.processed_events = file_manager.load_pickle("processed_events.pkl")
    
    def create_event(self, parsed_meeting: Dict, sender_email: str, sender_name: str) -> Optional[str]:
        """Create calendar event with enhanced duplicate detection"""
        if not parsed_meeting or not parsed_meeting.get("has_meeting"):
            return None
        
        # Validate required fields
        if not all([parsed_meeting.get("date"), parsed_meeting.get("start_time")]):
            logger.warning("Missing required date/time information")
            return None
        
        # Check for duplicates
        if self._is_duplicate_event(parsed_meeting, sender_email):
            logger.info(f"Duplicate event detected, skipping: {parsed_meeting.get('subject')}")
            return "duplicate_skipped"
        
        # Create event payload
        event_payload = self._build_event_payload(parsed_meeting, sender_email, sender_name)
        
        # Create the event
        create_event_url = "https://graph.microsoft.com/v1.0/me/events"
        response = requests.post(create_event_url, headers=self.headers, json=event_payload)
        
        if response.status_code == 201:
            event_data = response.json()
            
            # Save fingerprint to prevent future duplicates
            fingerprint = self._generate_event_fingerprint(parsed_meeting, sender_email)
            self.processed_events.add(fingerprint)
            self.file_manager.save_pickle(self.processed_events, "processed_events.pkl")
            
            logger.info(f"Event created successfully: {parsed_meeting.get('subject')}")
            return event_data["id"]
        else:
            logger.error(f"Failed to create event: {response.status_code} - {response.text}")
            return None
    
    def _is_duplicate_event(self, parsed_meeting: Dict, sender_email: str) -> bool:
        """Check if event is a duplicate"""
        fingerprint = self._generate_event_fingerprint(parsed_meeting, sender_email)
        return fingerprint in self.processed_events
    
    def _generate_event_fingerprint(self, parsed_meeting: Dict, sender_email: str) -> str:
        """Generate unique fingerprint for event"""
        subject = parsed_meeting.get("subject", "").lower().strip()
        date = parsed_meeting.get("date", "").strip()
        start_time = parsed_meeting.get("start_time", "").strip()
        
        fingerprint_data = f"{subject}|{date}|{start_time}|{sender_email.lower()}"
        return hashlib.md5(fingerprint_data.encode()).hexdigest()
    
    def _build_event_payload(self, parsed_meeting: Dict, sender_email: str, sender_name: str) -> Dict:
        """Build event payload for calendar creation"""
        date = parsed_meeting["date"]
        start_time = parsed_meeting["start_time"]
        end_time = parsed_meeting.get("end_time")
        
        # Calculate end time if not provided
        if not end_time:
            start_dt = datetime.strptime(f"{date} {start_time}", "%Y-%m-%d %H:%M")
            end_dt = start_dt + timedelta(hours=1)
            end_time = end_dt.strftime("%H:%M")
        
        # Build attendees list
        attendees_list = [{
            "emailAddress": {"address": sender_email, "name": sender_name},
            "type": "required"
        }]
        
        # Add other participants
        for participant in parsed_meeting.get("participants", []):
            if participant != sender_email:
                attendees_list.append({
                    "emailAddress": {"address": participant, "name": participant.split("@")[0]},
                    "type": "optional"
                })
        
        return {
            "subject": parsed_meeting.get("subject", "Meeting"),
            "start": {
                "dateTime": f"{date}T{start_time}:00.000",
                "timeZone": "Asia/Kolkata"
            },
            "end": {
                "dateTime": f"{date}T{end_time}:00.000",
                "timeZone": "Asia/Kolkata"
            },
            "location": {"displayName": parsed_meeting.get("location", "")},
            "attendees": attendees_list,
            "body": {
                "contentType": "Text",
                "content": f"Meeting Type: {parsed_meeting.get('meeting_type', 'N/A')}\n"
                          f"Action Items: {parsed_meeting.get('action_items', 'None')}\n"
                          f"Deadline: {parsed_meeting.get('deadline', 'N/A')}\n"
                          f"Priority: {parsed_meeting.get('priority', 'Medium')}"
            },
            "importance": "normal",
            "isReminderOn": True,
            "reminderMinutesBeforeStart": 15
        }

# ---------------- Authentication Manager ----------------
class AuthManager:
    """Enhanced authentication with better token management and error handling"""
    
import msal
import os

class AuthManager:
    def __init__(self, config: Config, file_manager: FileManager):
        self.config = config
        self.cache = msal.SerializableTokenCache()
        self.token_cache_file = "token_cache.bin"
        self._load_cache()

        self.app = msal.PublicClientApplication(
            client_id=self.config.CLIENT_ID,
            authority=self.config.AUTHORITY,
            token_cache=self.cache
        )

    def _load_cache(self):
        """Load token cache from file"""
        if os.path.exists(self.token_cache_file):
            with open(self.token_cache_file, "r") as f:
                self.cache.deserialize(f.read())

    def _save_cache(self):
        """Save token cache to file if changed"""
        if self.cache.has_state_changed:
            with open(self.token_cache_file, "w") as f:
                f.write(self.cache.serialize())

    def get_access_token(self) -> Optional[str]:
        """Get access token using silent flow and fallback to device code"""
        accounts = self.app.get_accounts()
        if accounts:
            result = self.app.acquire_token_silent(self.config.SCOPE, account=accounts[0])
            if result and "access_token" in result:
                self._save_cache()
                return result["access_token"]

        flow = self.app.initiate_device_flow(scopes=self.config.SCOPE)
        print(f"🌐 {flow['message']}")
        result = self.app.acquire_token_by_device_flow(flow)

        if "access_token" in result:
            self._save_cache()
            return result["access_token"]
        else:
            logger.error(f"Authentication failed: {result.get('error_description')}")
            return None

    
    def _device_flow_auth(self) -> Optional[str]:
        """Handle device flow authentication"""
        try:
            # Initiate device flow
            flow = self.app.initiate_device_flow(scopes=self.config.SCOPE)
            
            if "user_code" not in flow:
                logger.error("Failed to create device flow")
                return None
            
            # Display user instructions
            print("\n" + "="*60)
            print("🌐 MICROSOFT AUTHENTICATION REQUIRED")
            print("="*60)
            print(f"📱 Go to: {flow['verification_uri']}")
            print(f"🔑 Enter code: {flow['user_code']}")
            print("⏰ Waiting for authentication...")
            print("="*60)
            
            # Wait for user to complete authentication
            result = self.app.acquire_token_by_device_flow(flow)
            
            if "access_token" in result:
                logger.info("✅ Authentication successful!")
                return result["access_token"]
            else:
                error_msg = result.get('error_description', result.get('error', 'Unknown error'))
                logger.error(f"❌ Authentication failed: {error_msg}")
                
                # Common error handling
                if "AADSTS70016" in str(result):
                    logger.error("App registration issue. Please check your CLIENT_ID and TENANT_ID")
                elif "AADSTS650053" in str(result):
                    logger.error("App needs admin consent. Contact your Azure AD administrator")
                elif "AADSTS50020" in str(result):
                    logger.error("User account issue. Please check your Microsoft account")
                
                return None
                
        except Exception as e:
            logger.error(f"Device flow error: {str(e)}")
            return None
    
    def validate_token(self, token: str) -> bool:
        """Validate token by making a test request"""
        try:
            headers = {"Authorization": f"Bearer {token}"}
            response = requests.get(
                "https://graph.microsoft.com/v1.0/me",
                headers=headers,
                timeout=10
            )
            return response.status_code == 200
        except Exception as e:
            logger.error(f"Token validation failed: {e}")
            return False

# ---------------- Excel Report Generator ----------------
class ExcelReportGenerator:
    """Enhanced Excel report generation with better formatting"""
    
    def __init__(self, filename: str = "lead_tracking_data.xlsx"):
        self.filename = filename
        self.wb = Workbook()
        
    def generate_report(self, emails: List[EmailData], events: List[EventData], opportunities: Dict[str, OpportunityData]):
        """Generate comprehensive Excel report"""
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Create sheets
        self._create_opportunities_sheet(opportunities)
        self._create_interaction_log_sheet(emails, events, opportunities)
        self._create_summary_sheet(emails, events, opportunities)
        
        # Save workbook
        self.wb.save(self.filename)
        logger.info(f"📊 Report saved to {self.filename}")
    
    def _create_opportunities_sheet(self, opportunities: Dict[str, OpportunityData]):
        """Create opportunities master sheet with formatting"""
        ws = self.wb.create_sheet("Opportunities Master")
        
        # Headers
        headers = [
            "Opportunity ID", "Contact Name", "Company", "Email", "Phone",
            "Opportunity Title", "Lead Status", "Priority", "Source", "Notes", "Last Updated"
        ]
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for opp in opportunities.values():
            ws.append([
                opp.id, opp.contact_name, opp.company, opp.email, opp.phone,
                opp.opportunity_title, opp.lead_status, opp.priority, opp.source,
                opp.notes, opp.last_updated
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _create_interaction_log_sheet(self, emails: List[EmailData], events: List[EventData], opportunities: Dict[str, OpportunityData]):
        """Create interaction log sheet"""
        ws = self.wb.create_sheet("Interaction Log")
        
        # Headers
        headers = [
            "Opportunity ID", "Contact Name", "Interaction Date", "Interaction Type",
            "Summary", "Action Items", "Deadline", "Attendees", "Priority"
        ]
        ws.append(headers)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Add email interactions
        for email in emails:
            opp_id = self._generate_opportunity_id(email.sender_email, email.subject)
            opp = opportunities.get(opp_id)
            
            if email.parsed_meeting and email.parsed_meeting.get("has_meeting"):
                parsed = email.parsed_meeting
                ws.append([
                    opp_id,
                    opp.contact_name if opp else email.sender_name,
                    parsed.get("date", ""),
                    "Meeting (from Email)",
                    email.body_preview[:200] + "..." if len(email.body_preview) > 200 else email.body_preview,
                    parsed.get("action_items", ""),
                    parsed.get("deadline", ""),
                    ", ".join(parsed.get("participants", [])),
                    parsed.get("priority", "Medium")
                ])
        
        # Add calendar events
        for event in events:
            if event.attendees:
                attendee_emails = [a.strip() for a in event.attendees.split(",") if a.strip()]
                for attendee in attendee_emails:
                    for opp_id, opp in opportunities.items():
                        if opp.email == attendee:
                            ws.append([
                                opp_id,
                                opp.contact_name,
                                event.start[:10] if event.start else "",
                                "Calendar Meeting",
                                event.body[:200] + "..." if len(event.body) > 200 else event.body,
                                "",
                                "",
                                event.attendees,
                                "Medium"
                            ])
                            break
    
    def _create_summary_sheet(self, emails: List[EmailData], events: List[EventData], opportunities: Dict[str, OpportunityData]):
        """Create summary dashboard sheet"""
        ws = self.wb.create_sheet("Summary Dashboard")
        
        # Summary statistics
        ws.append(["📊 Lead Tracking Summary"])
        ws.append([])
        ws.append(["Total Emails Processed:", len(emails)])
        ws.append(["Total Calendar Events:", len(events)])
        ws.append(["Total Opportunities:", len(opportunities)])
        ws.append(["Meetings Detected:", sum(1 for e in emails if e.parsed_meeting and e.parsed_meeting.get("has_meeting"))])
        ws.append([])
        
        # Status breakdown
        status_counts = {}
        for opp in opportunities.values():
            status_counts[opp.lead_status] = status_counts.get(opp.lead_status, 0) + 1
        
        ws.append(["📈 Lead Status Breakdown:"])
        for status, count in status_counts.items():
            ws.append([f"  {status}:", count])
        
        # Format summary
        ws["A1"].font = Font(bold=True, size=16)
        ws["A8"].font = Font(bold=True, size=12)
    
    def _generate_opportunity_id(self, sender_email: str, subject: str) -> str:
        """Generate opportunity ID"""
        combined = f"{sender_email}_{subject}".lower()
        return hashlib.md5(combined.encode()).hexdigest()[:8]

# ---------------- Main Application ----------------
class LeadTracker:
    """Main application class with enhanced error handling"""
    
    def __init__(self):
        self.config = Config()
        self.file_manager = FileManager()
        self.auth_manager = AuthManager(self.config, self.file_manager)
        self.gemini_parser = GeminiParser(self.config.GEMINI_API_KEY)
        self.access_token = None
        self.headers = None
        
    def initialize(self) -> bool:
        """Initialize the application with comprehensive checks"""
        logger.info("🚀 Initializing Lead Tracker...")
        
        # Check configuration
        if not self._validate_config():
            return False
        
        # Get access token
        self.access_token = self.auth_manager.get_access_token()
        if not self.access_token:
            logger.error("❌ Failed to get access token")
            return False
        
        # Validate token
        if not self.auth_manager.validate_token(self.access_token):
            logger.error("❌ Token validation failed")
            return False
        
        self.headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json"
        }
        
        logger.info("✅ Initialization successful")
        return True
    
    def _validate_config(self) -> bool:
        """Validate configuration settings"""
        if not self.config.CLIENT_ID or self.config.CLIENT_ID == "your-client-id":
            logger.error("❌ CLIENT_ID not configured")
            return False
            
        if not self.config.TENANT_ID or self.config.TENANT_ID == "your-tenant-id":
            logger.error("❌ TENANT_ID not configured")
            return False
            
        if not self.config.GEMINI_API_KEY or self.config.GEMINI_API_KEY == "your-gemini-key":
            logger.error("❌ GEMINI_API_KEY not configured")
            return False
            
        logger.info("✅ Configuration validated")
        return True
    
    def run(self):
        """Main execution method with better error handling"""
        try:
            if not self.initialize():
                logger.error("❌ Initialization failed. Exiting.")
                return
            
            # Initialize managers
            calendar_manager = CalendarManager(self.headers, self.file_manager)
            
            # Process emails
            logger.info("📧 Processing emails...")
            emails = self._process_emails(calendar_manager)
            
            # Fetch calendar events
            logger.info("📅 Fetching calendar events...")
            events = self._fetch_calendar_events()
            
            # Generate opportunities
            logger.info("🎯 Generating opportunities...")
            opportunities = self._generate_opportunities(emails)
            
            # Create Excel report
            logger.info("📊 Creating Excel report...")
            report_generator = ExcelReportGenerator()
            report_generator.generate_report(emails, events, opportunities)
            
            # Log summary
            logger.info("="*60)
            logger.info("🎉 PROCESSING COMPLETE!")
            logger.info(f"📧 Emails processed: {len(emails)}")
            logger.info(f"📅 Calendar events: {len(events)}")
            logger.info(f"🎯 Opportunities: {len(opportunities)}")
            logger.info(f"🤝 Meetings detected: {sum(1 for e in emails if e.parsed_meeting and e.parsed_meeting.get('has_meeting'))}")
            logger.info("="*60)
            
        except Exception as e:
            logger.error(f"❌ Application error: {str(e)}")
            raise
    
    def _process_emails(self, calendar_manager: CalendarManager) -> List[EmailData]:
        """Process emails with meeting detection"""
        processed_emails = self.file_manager.load_pickle("processed_emails.pkl")
        emails = []
        
        try:
            email_url = "https://graph.microsoft.com/v1.0/me/messages?$top=50&$orderby=receivedDateTime desc"
            response = requests.get(email_url, headers=self.headers, timeout=30)
            
            if response.status_code != 200:
                logger.error(f"❌ Failed to fetch emails: {response.status_code} - {response.text}")
                return emails
            
            messages = response.json().get("value", [])
            logger.info(f"📧 Found {len(messages)} emails")
            
            for i, msg in enumerate(messages):
                email_id = msg["id"]
                
                # Skip if already processed
                if email_id in processed_emails:
                    continue
                
                logger.info(f"📧 Processing email {i+1}/{len(messages)}: {msg.get('subject', 'No Subject')}")
                
                # Create email data object
                email_data = EmailData(
                    id=email_id,
                    sender_name=msg.get("sender", {}).get("emailAddress", {}).get("name", "Unknown"),
                    sender_email=msg.get("sender", {}).get("emailAddress", {}).get("address", "unknown@example.com"),
                    subject=msg.get("subject", "No Subject"),
                    received=msg.get("receivedDateTime", ""),
                    body_preview=msg.get("bodyPreview", "")
                )
                
                # Parse email with Gemini
                parsed = self.gemini_parser.parse_email_with_retry(
                    email_data.body_preview, 
                    email_data.subject
                )
                email_data.parsed_meeting = parsed
                
                # Create calendar event if meeting detected
                if parsed and parsed.get("has_meeting"):
                    logger.info(f"🤝 Meeting detected: {email_data.subject}")
                    event_id = calendar_manager.create_event(
                        parsed, 
                        email_data.sender_email, 
                        email_data.sender_name
                    )
                    email_data.calendar_event_id = event_id
                
                emails.append(email_data)
                processed_emails.add(email_id)
            
            # Save processed emails
            self.file_manager.save_pickle(processed_emails, "processed_emails.pkl")
            
        except Exception as e:
            logger.error(f"❌ Error processing emails: {str(e)}")
        
        return emails
    
    def _fetch_calendar_events(self) -> List[EventData]:
        """Fetch calendar events with better error handling"""
        events = []
        
        try:
            start_time = datetime.now(timezone.utc) - timedelta(days=7)
            end_time = datetime.now(timezone.utc) + timedelta(days=7)
            
            params = {
                "startDateTime": start_time.isoformat(),
                "endDateTime": end_time.isoformat()
            }
            
            calendar_url = "https://graph.microsoft.com/v1.0/me/calendarView"
            response = requests.get(calendar_url, headers=self.headers, params=params, timeout=30)
            
            if response.status_code == 200:
                calendar_events = response.json().get("value", [])
                logger.info(f"📅 Found {len(calendar_events)} calendar events")
                
                for event in calendar_events:
                    attendees = ", ".join([
                        a["emailAddress"]["address"] 
                        for a in event.get("attendees", [])
                    ])
                    
                    events.append(EventData(
                        id=event["id"],
                        subject=event.get("subject", ""),
                        start=event.get("start", {}).get("dateTime", ""),
                        end=event.get("end", {}).get("dateTime", ""),
                        location=event.get("location", {}).get("displayName", ""),
                        attendees=attendees,
                        body=event.get("bodyPreview", "")
                    ))
            else:
                logger.error(f"❌ Failed to fetch calendar events: {response.status_code}")
                
        except Exception as e:
            logger.error(f"❌ Error fetching calendar events: {str(e)}")
        
        return events
    
    def _generate_opportunities(self, emails: List[EmailData]) -> Dict[str, OpportunityData]:
        """Generate opportunities from emails with better data extraction"""
        opportunities = {}
        
        for email in emails:
            try:
                opp_id = self._generate_opportunity_id(email.sender_email, email.subject)
                
                # Extract company from email domain
                company = email.sender_email.split("@")[1].split(".")[0].title()
                
                # Determine priority based on meeting detection
                priority = "High" if email.parsed_meeting and email.parsed_meeting.get("has_meeting") else "Medium"
                
                # Determine lead status
                lead_status = "Meeting Scheduled" if email.parsed_meeting and email.parsed_meeting.get("has_meeting") else "New"
                
                opportunities[opp_id] = OpportunityData(
                    id=opp_id,
                    contact_name=email.sender_name,
                    company=company,
                    email=email.sender_email,
                    phone="",
                    opportunity_title=email.subject,
                    lead_status=lead_status,
                    notes=email.body_preview[:200] + "..." if len(email.body_preview) > 200 else email.body_preview,
                    last_updated=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    priority=priority,
                    source="Email"
                )
                
            except Exception as e:
                logger.error(f"❌ Error generating opportunity for {email.sender_email}: {str(e)}")
                continue
        
        return opportunities
    
    def _generate_opportunity_id(self, sender_email: str, subject: str) -> str:
        """Generate opportunity ID"""
        combined = f"{sender_email}_{subject}".lower()
        return hashlib.md5(combined.encode()).hexdigest()[:8]

# ---------------- Diagnostics and Troubleshooting ----------------
class DiagnosticsTool:
    """Diagnostic tool for troubleshooting authentication issues"""
    
    def __init__(self, config: Config):
        self.config = config
    
    def run_diagnostics(self):
        """Run comprehensive diagnostics"""
        print("\n" + "="*60)
        print("🔍 LEAD TRACKER DIAGNOSTICS")
        print("="*60)
        
        # Check configuration
        self._check_config()
        
        # Check network connectivity
        self._check_connectivity()
        
        # Check MSAL setup
        self._check_msal_setup()
        
        print("="*60)
        print("✅ Diagnostics complete")
        print("="*60)
    
    def _check_config(self):
        """Check configuration settings"""
        print("\n📋 Configuration Check:")
        print(f"  CLIENT_ID: {'✅ Set' if self.config.CLIENT_ID and self.config.CLIENT_ID != 'your-client-id' else '❌ Not set'}")
        print(f"  TENANT_ID: {'✅ Set' if self.config.TENANT_ID and self.config.TENANT_ID != 'your-tenant-id' else '❌ Not set'}")
        print(f"  GEMINI_API_KEY: {'✅ Set' if self.config.GEMINI_API_KEY and self.config.GEMINI_API_KEY != 'your-gemini-key' else '❌ Not set'}")
        print(f"  Authority URL: {self.config.AUTHORITY}")
        print(f"  Scopes: {', '.join(self.config.SCOPE)}")
    
    def _check_connectivity(self):
        """Check network connectivity"""
        print("\n🌐 Network Connectivity Check:")
        
        urls_to_test = [
            ("Microsoft Graph", "https://graph.microsoft.com/v1.0/"),
            ("Microsoft Login", "https://login.microsoftonline.com/"),
            ("Google AI", "https://generativelanguage.googleapis.com/")
        ]
        
        for name, url in urls_to_test:
            try:
                response = requests.get(url, timeout=10)
                print(f"  {name}: ✅ Connected ({response.status_code})")
            except Exception as e:
                print(f"  {name}: ❌ Failed ({str(e)})")
    
    def _check_msal_setup(self):
        """Check MSAL setup"""
        print("\n🔐 MSAL Setup Check:")
        
        try:
            app = msal.PublicClientApplication(
                client_id=self.config.CLIENT_ID,
                authority=self.config.AUTHORITY
            )
            print("  MSAL App Creation: ✅ Success")
            
            # Check for existing accounts
            accounts = app.get_accounts()
            print(f"  Cached Accounts: {len(accounts)} found")
            
            if accounts:
                for i, account in enumerate(accounts):
                    print(f"    Account {i+1}: {account.get('username', 'Unknown')}")
            
        except Exception as e:
            print(f"  MSAL Setup: ❌ Failed ({str(e)})")

# ---------------- Main Execution with Enhanced Error Handling ----------------
def main():
    """Main execution function with comprehensive error handling"""
    try:
        # Check if diagnostics mode is requested
        if len(os.sys.argv) > 1 and os.sys.argv[1] == "--diagnostics":
            config = Config()
            diagnostics = DiagnosticsTool(config)
            diagnostics.run_diagnostics()
            return
        
        # Run main application
        logger.info("🚀 Starting Lead Tracker Application")
        app = LeadTracker()
        app.run()
        
    except KeyboardInterrupt:
        logger.info("⚠️  Application interrupted by user")
    except Exception as e:
        logger.error(f"❌ Critical application error: {str(e)}")
        logger.error("💡 Try running with --diagnostics flag for troubleshooting")
        raise

if __name__ == "__main__":
    main()